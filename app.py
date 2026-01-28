"""
Streamlit App: ASUS Credit Note PDF Extractor
Giao diện tuần tự với quản lý người dùng
"""

import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Page config
st.set_page_config(
    page_title="ASUS Credit Note Extractor",
    page_icon="📄",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #2E75B6;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.1rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .step-header {
        font-size: 1.5rem;
        font-weight: bold;
        color: #2E75B6;
        padding: 0.5rem 0;
        border-bottom: 2px solid #2E75B6;
        margin-bottom: 1rem;
    }
    .step-box {
        padding: 1.5rem;
        background-color: #f8f9fa;
        border-radius: 0.5rem;
        border: 1px solid #dee2e6;
        margin-bottom: 1rem;
    }
    .step-number {
        display: inline-block;
        width: 35px;
        height: 35px;
        background-color: #2E75B6;
        color: white;
        border-radius: 50%;
        text-align: center;
        line-height: 35px;
        font-weight: bold;
        margin-right: 10px;
    }
    .step-active {
        background-color: #2E75B6;
    }
    .step-completed {
        background-color: #28A745;
    }
    .step-pending {
        background-color: #ddd;
        color: #666;
    }
    .user-badge {
        background-color: #2E75B6;
        color: white;
        padding: 0.3rem 0.8rem;
        border-radius: 1rem;
        font-size: 0.9rem;
    }
    .stats-card {
        text-align: center;
        padding: 1rem;
        background: linear-gradient(135deg, #2E75B6, #1a4a7a);
        color: white;
        border-radius: 0.5rem;
    }
    .stats-number {
        font-size: 2rem;
        font-weight: bold;
    }
    .stats-label {
        font-size: 0.9rem;
        opacity: 0.9;
    }
</style>
""", unsafe_allow_html=True)


# ==================== SESSION STATE ====================

if 'current_step' not in st.session_state:
    st.session_state.current_step = 1

if 'user_name' not in st.session_state:
    st.session_state.user_name = ""

if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []

if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None

if 'processing_log' not in st.session_state:
    st.session_state.processing_log = []


# ==================== EXTRACTION FUNCTIONS ====================

def pdf_to_text(pdf_file) -> str:
    """Đọc file PDF và trả về text content."""
    text_content = []
    try:
        with pdfplumber.open(pdf_file) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                page_text = page.extract_text()
                if page_text:
                    text_content.append(f"=== PAGE {page_num} ===")
                    text_content.append(page_text)
                    text_content.append("")
    except Exception as e:
        return ""
    return "\n".join(text_content)


def extract_cn_no(text: str) -> str:
    match = re.search(r'CN NO\s*:\s*(\d+)', text)
    return match.group(1) if match else ""


def extract_product_line(text: str) -> str:
    match = re.search(r'Credit Note Remark:\s*(.+?)(?:\r?\n|$)', text)
    return match.group(1).strip() if match else ""


def extract_total(text: str) -> str:
    match = re.search(r'Total:\s*([\d,]+\.?\d*)', text)
    return match.group(1) if match else ""


def extract_items(text: str) -> list:
    """Trích xuất các item từ text."""
    items = []
    lines = text.split('\n')
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        item_match = re.match(
            r'^(\d+\.\d+)\s+([A-Z0-9\-]+)\s+(\d+)\s+(.+)$',
            line
        )
        
        if item_match:
            no = item_match.group(1)
            part_no = item_match.group(2)
            rest = item_match.group(4).strip()
            
            numbers = re.findall(r'[\d,]+\.?\d*', rest)
            fob = numbers[-1] if numbers else ""
            
            product = ""
            for j in range(i + 1, min(i + 3, len(lines))):
                next_line = lines[j].strip()
                if next_line.startswith(('Model:', '===', 'No ', 'Page', 'SO:', 'Note:')):
                    continue
                if next_line.startswith('AS '):
                    product = next_line
                    break
                elif '/' in next_line and not next_line.startswith(('EAN', 'MODEL')):
                    product = next_line
                    break
            
            serial = ""
            for j in range(i + 1, min(i + 15, len(lines))):
                note_line = lines[j].strip()
                if note_line.startswith(('===', 'Page:', 'CN#', 'No Description', 
                                         'ASUS GLOBAL', '10 Changi', 'Reg. No',
                                         'Credit Note', 'To :', 'Address', 'Attn',
                                         'Fax', 'Date', 'CN Reason', 'Credit Note Remark')):
                    continue
                if re.match(r'^\d+\.\d+\s+[A-Z0-9\-]+\s+\d+\s+', note_line):
                    break
                if 'SN:' in note_line:
                    sn_match = re.search(r'SN:([A-Z0-9]+)', note_line)
                    if sn_match:
                        serial = sn_match.group(1)
                    else:
                        memo_match = re.search(r'MEMO:([A-Z0-9]+)', note_line)
                        if memo_match:
                            serial = memo_match.group(1)
                    break
            
            invoice = ""
            for j in range(i + 1, min(i + 10, len(lines))):
                inv_line = lines[j].strip()
                if re.match(r'^\d+\.\d+\s+[A-Z0-9\-]+\s+\d+\s+', inv_line):
                    break
                inv_match = re.search(r'INVOICE[:\s]+([\d]+)', inv_line)
                if inv_match:
                    invoice = inv_match.group(1)
                    break
            
            items.append({
                'No': no,
                'Part No': part_no,
                'Product': product,
                'Serial': serial,
                'FOB': fob,
                'Invoice': invoice
            })
        
        i += 1
    
    return items


def parse_rebate_files(pdf_texts: dict) -> dict:
    rebate_mapping = {}
    for filename, content in pdf_texts.items():
        if 'REBATE FOR INVOICE:' not in content:
            continue
        cn_no = extract_cn_no(content)
        rebate_pattern = re.findall(
            r'REBATE FOR INVOICE:\s*(\d+)\s+([\d,.]+)',
            content
        )
        for invoice, amount in rebate_pattern:
            rebate_mapping[invoice] = {
                'CN_Landing': cn_no,
                'Landing_cost': amount
            }
    return rebate_mapping


def process_pdf_text(filename: str, text: str, rebate_mapping: dict) -> list:
    if 'REBATE FOR INVOICE:' in text:
        return []
    
    pdf_name = filename
    cn_fob = extract_cn_no(text)
    product_line = extract_product_line(text)
    items = extract_items(text)
    total = extract_total(text)
    
    file_landing_costs = []
    file_cn_landing = ''
    
    for item in items:
        invoice = item.get('Invoice', '')
        if invoice and invoice in rebate_mapping:
            cn_landing = rebate_mapping[invoice]['CN_Landing']
            landing_cost = rebate_mapping[invoice]['Landing_cost']
            file_landing_costs.append(landing_cost)
            if cn_landing:
                file_cn_landing = cn_landing
    
    records = []
    for item in items:
        records.append({
            'Tên file PDF': pdf_name,
            'Product': item['Product'],
            'Product line': product_line,
            'Serial': item['Serial'],
            'Part No': item['Part No'],
            'FOB': item['FOB'],
            'CN FOB': cn_fob,
            'CN Landing': file_cn_landing,
            'Landing cost': ''
        })
    
    total_landing_cost = ''
    if file_landing_costs:
        try:
            total_landing_cost = sum(float(lc.replace(',', '')) for lc in file_landing_costs)
            total_landing_cost = f"{total_landing_cost:.2f}"
        except:
            total_landing_cost = ', '.join(file_landing_costs)
    
    if records:
        records.append({
            'Tên file PDF': pdf_name,
            'Product': '',
            'Product line': '',
            'Serial': '',
            'Part No': 'TOTAL',
            'FOB': total,
            'CN FOB': cn_fob,
            'CN Landing': file_cn_landing,
            'Landing cost': total_landing_cost
        })
    
    return records


def create_excel_with_formatting(df: pd.DataFrame) -> io.BytesIO:
    output = io.BytesIO()
    df.to_excel(output, index=False, sheet_name='Data', engine='openpyxl')
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb.active
    
    merge_columns = {'Tên file PDF': 1, 'CN FOB': 7, 'CN Landing': 8}
    current_file = None
    start_row = 2
    
    for idx, row in df.iterrows():
        excel_row = idx + 2
        file_name = row['Tên file PDF']
        is_total = row['Part No'] == 'TOTAL'
        
        if current_file is None:
            current_file = file_name
            start_row = excel_row
        elif is_total:
            end_row = excel_row
            for col_name, col_idx in merge_columns.items():
                if end_row > start_row:
                    col_letter = get_column_letter(col_idx)
                    ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
                    ws[f'{col_letter}{start_row}'].alignment = Alignment(vertical='center')
            current_file = None
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row in range(2, ws.max_row + 1):
        is_total_row = ws.cell(row=row, column=5).value == 'TOTAL'
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if is_total_row:
                cell.font = total_font
                cell.fill = total_fill
            if col in [5, 6, 7, 8, 9]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    column_widths = {'A': 25, 'B': 45, 'C': 25, 'D': 18, 'E': 18, 'F': 12, 'G': 18, 'H': 18, 'I': 15}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    ws.freeze_panes = 'A2'
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def log_activity(user: str, action: str, details: str = ""):
    """Ghi log hoạt động người dùng."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = {
        'timestamp': timestamp,
        'user': user,
        'action': action,
        'details': details
    }
    st.session_state.processing_log.append(log_entry)
    # In log ra console (sẽ hiển thị trong Streamlit Cloud logs)
    print(f"[{timestamp}] User: {user} | Action: {action} | {details}")


# ==================== STEP INDICATOR ====================

def render_step_indicator():
    """Hiển thị thanh tiến trình các bước."""
    steps = ["Đăng nhập", "Upload PDF", "Xử lý", "Tải xuống"]
    
    cols = st.columns(len(steps))
    for i, (col, step_name) in enumerate(zip(cols, steps)):
        step_num = i + 1
        with col:
            if step_num < st.session_state.current_step:
                status_class = "step-completed"
                icon = "✓"
            elif step_num == st.session_state.current_step:
                status_class = "step-active"
                icon = str(step_num)
            else:
                status_class = "step-pending"
                icon = str(step_num)
            
            st.markdown(f"""
            <div style="text-align: center;">
                <span class="step-number {status_class}">{icon}</span>
                <div style="font-weight: {'bold' if step_num == st.session_state.current_step else 'normal'}; 
                            color: {'#2E75B6' if step_num <= st.session_state.current_step else '#999'};">
                    {step_name}
                </div>
            </div>
            """, unsafe_allow_html=True)


# ==================== MAIN UI ====================

# Header
st.markdown('<div class="main-header">📄 ASUS Credit Note Extractor</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Trích xuất dữ liệu từ PDF Credit Note sang Excel</div>', unsafe_allow_html=True)

# User badge (nếu đã đăng nhập)
if st.session_state.user_name:
    st.markdown(f"""
    <div style="text-align: right; margin-bottom: 1rem;">
        <span class="user-badge">👤 {st.session_state.user_name}</span>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# Step indicator
render_step_indicator()

st.markdown("---")

# ==================== STEP 1: LOGIN ====================

if st.session_state.current_step == 1:
    st.markdown('<div class="step-header">📝 Bước 1: Nhập thông tin người dùng</div>', unsafe_allow_html=True)
    
    with st.container():
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown('<div class="step-box">', unsafe_allow_html=True)
            
            user_name = st.text_input(
                "Họ và tên",
                placeholder="Nhập họ và tên của bạn...",
                help="Tên của bạn sẽ được ghi nhận để quản lý"
            )
            
            if st.button("✅ Tiếp tục", type="primary", use_container_width=True):
                if user_name.strip():
                    st.session_state.user_name = user_name.strip()
                    st.session_state.current_step = 2
                    log_activity(user_name.strip(), "LOGIN", "User logged in")
                    st.rerun()
                else:
                    st.error("⚠️ Vui lòng nhập họ và tên!")
            
            st.markdown('</div>', unsafe_allow_html=True)


# ==================== STEP 2: UPLOAD ====================

elif st.session_state.current_step == 2:
    st.markdown('<div class="step-header">📁 Bước 2: Upload các file PDF</div>', unsafe_allow_html=True)
    
    uploaded_files = st.file_uploader(
        "Kéo thả hoặc chọn các file PDF Credit Note",
        type=['pdf'],
        accept_multiple_files=True,
        help="Bạn có thể upload nhiều file cùng lúc"
    )
    
    if uploaded_files:
        st.success(f"✅ Đã chọn {len(uploaded_files)} file")
        
        # Preview file names
        with st.expander("📋 Xem danh sách file", expanded=True):
            for i, f in enumerate(uploaded_files, 1):
                st.text(f"{i}. {f.name}")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("⬅️ Quay lại", use_container_width=True):
            st.session_state.current_step = 1
            st.rerun()
    with col2:
        if st.button("🚀 Xử lý file", type="primary", use_container_width=True, disabled=not uploaded_files):
            if uploaded_files:
                st.session_state.uploaded_files = uploaded_files
                st.session_state.current_step = 3
                log_activity(st.session_state.user_name, "UPLOAD", f"Uploaded {len(uploaded_files)} files")
                st.rerun()


# ==================== STEP 3: PROCESS ====================

elif st.session_state.current_step == 3:
    st.markdown('<div class="step-header">⚙️ Bước 3: Xử lý dữ liệu</div>', unsafe_allow_html=True)
    
    uploaded_files = st.session_state.uploaded_files
    
    # Progress
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # Process PDFs
    pdf_texts = {}
    
    for i, uploaded_file in enumerate(uploaded_files):
        status_text.text(f"📖 Đang đọc: {uploaded_file.name}...")
        progress_bar.progress((i + 1) / len(uploaded_files) * 0.4)
        
        uploaded_file.seek(0)
        text = pdf_to_text(uploaded_file)
        if text:
            pdf_texts[uploaded_file.name] = text
    
    # Parse REBATE
    status_text.text("🔍 Đang phân tích REBATE files...")
    progress_bar.progress(0.5)
    rebate_mapping = parse_rebate_files(pdf_texts)
    
    # Extract data
    all_records = []
    rebate_count = 0
    processed_files = []
    
    for i, (filename, text) in enumerate(pdf_texts.items()):
        status_text.text(f"📊 Đang xử lý: {filename}...")
        progress_bar.progress(0.5 + (i + 1) / len(pdf_texts) * 0.5)
        
        records = process_pdf_text(filename, text, rebate_mapping)
        if records:
            all_records.extend(records)
            processed_files.append(filename)
        else:
            rebate_count += 1
    
    progress_bar.progress(1.0)
    status_text.text("✅ Hoàn thành xử lý!")
    
    if all_records:
        df = pd.DataFrame(all_records)
        columns_order = [
            'Tên file PDF', 'Product', 'Product line', 
            'Serial', 'Part No', 'FOB', 'CN FOB',
            'CN Landing', 'Landing cost'
        ]
        df = df[columns_order]
        
        st.session_state.processed_data = df
        
        log_activity(
            st.session_state.user_name, 
            "PROCESS", 
            f"Processed {len(processed_files)} files, {len(all_records)} records"
        )
        
        # Statistics
        st.markdown("### 📊 Thống kê")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div class="stats-card">
                <div class="stats-number">{len(uploaded_files)}</div>
                <div class="stats-label">File upload</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="stats-card">
                <div class="stats-number">{len(processed_files)}</div>
                <div class="stats-label">File xử lý</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="stats-card">
                <div class="stats-number">{rebate_count}</div>
                <div class="stats-label">REBATE files</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div class="stats-card">
                <div class="stats-number">{len(all_records)}</div>
                <div class="stats-label">Records</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("### 👀 Xem trước dữ liệu")
        st.dataframe(df, use_container_width=True, height=300)
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("⬅️ Upload lại", use_container_width=True):
                st.session_state.current_step = 2
                st.session_state.processed_data = None
                st.rerun()
        with col2:
            if st.button("📥 Tiếp tục tải xuống", type="primary", use_container_width=True):
                st.session_state.current_step = 4
                st.rerun()
    else:
        st.error("⚠️ Không trích xuất được dữ liệu. Vui lòng kiểm tra file PDF.")
        if st.button("⬅️ Quay lại", use_container_width=True):
            st.session_state.current_step = 2
            st.rerun()


# ==================== STEP 4: DOWNLOAD ====================

elif st.session_state.current_step == 4:
    st.markdown('<div class="step-header">📥 Bước 4: Tải xuống kết quả</div>', unsafe_allow_html=True)
    
    df = st.session_state.processed_data
    
    if df is not None:
        st.success(f"✅ Đã xử lý thành công {len(df)} records!")
        
        # Generate Excel
        excel_file = create_excel_with_formatting(df)
        
        # Download info
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown('<div class="step-box" style="text-align: center;">', unsafe_allow_html=True)
            
            st.markdown("### 📄 File sẵn sàng tải xuống")
            st.markdown(f"**Người thực hiện:** {st.session_state.user_name}")
            st.markdown(f"**Thời gian:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            st.markdown(f"**Số records:** {len(df)}")
            
            st.markdown("---")
            
            # Download button
            download_clicked = st.download_button(
                label="📥 Tải xuống Excel",
                data=excel_file,
                file_name=f"extracted_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            
            if download_clicked:
                log_activity(st.session_state.user_name, "DOWNLOAD", f"Downloaded {len(df)} records")
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Xử lý file mới", use_container_width=True):
                st.session_state.current_step = 2
                st.session_state.uploaded_files = []
                st.session_state.processed_data = None
                st.rerun()
        with col2:
            if st.button("🚪 Đăng xuất", use_container_width=True):
                log_activity(st.session_state.user_name, "LOGOUT", "User logged out")
                st.session_state.current_step = 1
                st.session_state.user_name = ""
                st.session_state.uploaded_files = []
                st.session_state.processed_data = None
                st.rerun()
    else:
        st.error("⚠️ Không có dữ liệu. Vui lòng quay lại bước xử lý.")
        if st.button("⬅️ Quay lại", use_container_width=True):
            st.session_state.current_step = 3
            st.rerun()


# Footer
st.markdown("---")
st.markdown(
    f"<div style='text-align: center; color: #888; font-size: 0.9rem;'>Made with ❤️ by Trần Duy Quang | {datetime.now().strftime('%Y-%m-%d')}</div>",
    unsafe_allow_html=True
)
